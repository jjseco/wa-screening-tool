import pandas as pd
import geopandas as gpd
import os
import time
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import OUTPUTS_TABLES, ANALYSIS_CRS
from scripts.registry import LAYER_REGISTRY


def _parse_dataset(source):
    """Extract dataset code from a source string such as 'Official WA - DWER-046'."""
    if " - " in source:
        return source.split(" - ", 1)[1]
    return source


def _get_provider(source):
    """Extract provider name from source string such as 'Official WA - DWER-046'."""
    if " - " in source:
        return source.split(" - ", 1)[0]
    return source


def _get_citation(layer_key, entry):
    """Build a formatted citation string for a registry layer."""
    label = entry.get("label", layer_key)
    code = entry.get("dataset_code", "")
    source = entry.get("source", "")
    dataset_year = entry.get("dataset_year", 2024)
    accessed_year = entry.get("accessed_year", 2026)

    if "OSM" in source:
        return (
            f"OpenStreetMap contributors ({dataset_year}). {label}. "
            f"Geofabrik Australia extract. Accessed March {accessed_year} from download.geofabrik.de"
        )
    provider = _get_provider(source) if source else ""
    if provider:
        return (
            f"{provider} ({dataset_year}). {label} [{code}]. "
            f"Accessed March {accessed_year} from data.wa.gov.au"
        )
    return ""


def get_result_row(r, buffer_distances):
    """Convert a result dict to a structured row."""
    nearest = r["nearest_distance_m"]
    nearest_str = "—" if r["site_intersect"] else (str(nearest) if nearest else "—")

    nearest_name = ""
    if r["nearby_names"]:
        nearest_name = r["nearby_names"][0]

    detected = ", ".join(r["nearby_names"]) if r["nearby_names"] else "—"

    row = {
        "Theme": r["theme"],
        "Layer": r["layer"],
        "Data Source": (lambda e: f"{e.get('label', r['layer'])} ({e.get('dataset_code', '')})")(LAYER_REGISTRY.get(r["layer_key"], {})),
        "Present": r["present"],
        "Max Relevance": r["max_relevance"],
        "Primary Trigger": r["primary_trigger"],
        "Site Intersect": r["site_intersect"],
    }
    for d in buffer_distances:
        row[f"Within {d}m"] = r.get(f"within_{d}m", False)
    for d in buffer_distances:
        row[f"Count {d}m"] = r.get(f"count_{d}m", 0)
    row["Nearest Distance (m)"] = nearest_str
    row["Nearest Feature"] = nearest_name
    row["Detected Features"] = detected
    row["Comment"] = r["comment"]
    row["Interpretation"] = r.get("interpretation", "")
    return row


def format_sheet(ws, df):
    """Apply formatting to a worksheet."""
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_num, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Freeze header and first 3 columns
    ws.freeze_panes = "D2"

    # Auto column width
    for col_num, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_num)
        max_length = max(
            len(str(col_name)),
            max((len(str(ws.cell(row=r, column=col_num).value or ""))
                 for r in range(2, ws.max_row + 1)), default=0)
        )
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

    # Wrap text for detected features and comment columns
    wrap_cols = ["Detected Features", "Comment", "Primary Trigger", "Interpretation"]
    for col_num, col_name in enumerate(df.columns, 1):
        if col_name in wrap_cols:
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 35
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_num).alignment = Alignment(wrap_text=True)

    # Row height
    ws.row_dimensions[1].height = 30


def export_to_excel(results, site_name, site_gdf=None, layers=None, buffers=None, theme_summary=None, session_id=""):
    """Export screening results to a structured Excel file with one sheet per theme."""

    out_dir = "/tmp/wa_outputs/tables"
    os.makedirs(out_dir, exist_ok=True)
    prefix = f"{session_id}_" if session_id else ""
    file_name = f"{prefix}{site_name.replace(' ', '_')}_screening_results.xlsx"
    file_path = os.path.join(out_dir, file_name)

    buffer_distances = sorted(buffers.keys()) if buffers else [100, 250, 500]

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:

        # --------------------------------------------------------
        # SHEET 1 — RISK SUMMARY
        # --------------------------------------------------------
        if theme_summary:
            rating_display = {"HIGH": "🔴 HIGH", "MEDIUM": "🟡 MEDIUM", "LOW": "🟢 LOW", "NONE": "⚪ NONE"}
            risk_rows = []
            for theme, ts in theme_summary.items():
                risk_rows.append({
                    "Theme": theme,
                    "Risk Rating": rating_display.get(ts["rating"], ts["rating"]),
                    "Score": ts["score"],
                    "Key Finding": ts["key_finding"],
                    "Interpretation": ts["interpretation"],
                })
            df_risk = pd.DataFrame(risk_rows)
            df_risk.to_excel(writer, sheet_name="Risk Summary", index=False)
            format_sheet(writer.sheets["Risk Summary"], df_risk)

        # --------------------------------------------------------
        # SHEET 2 — SUMMARY
        # --------------------------------------------------------
        summary_rows = []
        themes = list(dict.fromkeys(r["theme"] for r in results))

        for theme in themes:
            theme_results = [r for r in results if r["theme"] == theme]
            for r in theme_results:
                summary_rows.append(get_result_row(r, buffer_distances))
            summary_rows.append({col: "" for col in summary_rows[0].keys()})

        df_summary = pd.DataFrame(summary_rows)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        format_sheet(writer.sheets["Summary"], df_summary)

        # --------------------------------------------------------
        # SHEETS PER THEME
        # --------------------------------------------------------
        sheet_name_map = {
            "Ecology": "Ecology",
            "ASS / Soils / Geology": "ASS - Soils",
            "Heritage": "Heritage",
            "Groundwater / Hydrology": "Groundwater",
            "Contaminated Land": "Contaminated Land",
            "Planning / Local Government": "Planning",
            "Sensitive Receptors": "Sensitive Receptors",
        }

        for theme in themes:
            theme_results = [r for r in results if r["theme"] == theme]
            if not theme_results:
                continue
            rows = [get_result_row(r, buffer_distances) for r in theme_results]
            sheet = sheet_name_map.get(theme, theme[:31])
            df_theme = pd.DataFrame(rows)
            df_theme.to_excel(writer, sheet_name=sheet, index=False)
            format_sheet(writer.sheets[sheet], df_theme)

        # --------------------------------------------------------
        # SHEET — PARCELS DETAIL
        # --------------------------------------------------------
        if site_gdf is not None and layers is not None and buffers is not None:
            if "cadastre" in layers:
                try:
                    cadastre_gdf = layers["cadastre"]
                    site_point = site_gdf.geometry.iloc[0]
                    parcel_rows = []
                    seen = set()
                    smallest_dist = buffer_distances[0]

                    for dist in buffer_distances:
                        buf_union = buffers[dist].geometry.union_all()
                        nearby = cadastre_gdf[cadastre_gdf.geometry.intersects(buf_union)].copy()
                        for idx, row in nearby.iterrows():
                            if idx in seen:
                                continue
                            seen.add(idx)
                            centroid = row.geometry.centroid
                            raw_distance = float(row.geometry.distance(site_point))
                            distance = "Intersects site" if raw_distance < 1 else round(raw_distance, 1)
                            point_gdf = gpd.GeoDataFrame(
                                geometry=[centroid], crs=ANALYSIS_CRS
                            ).to_crs("EPSG:4326")
                            lat = round(point_gdf.geometry.iloc[0].y, 6)
                            lon = round(point_gdf.geometry.iloc[0].x, 6)
                            parcel_rows.append({
                                "Buffer Zone": f"{dist}m",
                                "Distance to Site (m)": distance,
                                "Latitude": lat,
                                "Longitude": lon,
                                "Address": "",
                            })

                    if parcel_rows:
                        # Reverse geocode parcels in the innermost buffer only
                        try:
                            from geopy.geocoders import Nominatim
                            geolocator = Nominatim(user_agent="wa_screening_tool_export")
                            for prow in parcel_rows:
                                if prow["Buffer Zone"] == f"{smallest_dist}m":
                                    try:
                                        loc = geolocator.reverse(
                                            f"{prow['Latitude']}, {prow['Longitude']}",
                                            language="en",
                                        )
                                        if loc:
                                            prow["Address"] = loc.address
                                    except Exception:
                                        pass
                                    time.sleep(1)
                        except ImportError:
                            pass

                        df_parcels = pd.DataFrame(parcel_rows)
                        df_parcels.to_excel(writer, sheet_name="Parcels Detail", index=False)
                        ws_parcels = writer.sheets["Parcels Detail"]
                        format_sheet(ws_parcels, df_parcels)
                        note_row = ws_parcels.max_row + 2
                        ws_parcels.cell(row=note_row, column=1).value = (
                            "Addresses shown for inner buffer only. Source: OpenStreetMap Nominatim"
                        )

                except Exception as e:
                    print(f"Parcel detail export error: {e}")

        # --------------------------------------------------------
        # SHEET — RECEPTORS DETAIL
        # --------------------------------------------------------
        if site_gdf is not None and layers is not None and buffers is not None:
            receptor_layers = {
                "hospitals": "Hospital",
                "schools": "School",
                "community_facilities": "Community Facility",
            }

            receptor_rows = []
            site_point = site_gdf.geometry.iloc[0]

            for layer_key, receptor_type in receptor_layers.items():
                if layer_key not in layers:
                    continue
                try:
                    layer_gdf = layers[layer_key]
                    buf_union = buffers[max(buffer_distances)].geometry.union_all()
                    nearby = layer_gdf[layer_gdf.geometry.intersects(buf_union)].copy()

                    for _, row in nearby.iterrows():
                        centroid = row.geometry.centroid
                        raw_distance = float(row.geometry.distance(site_point))
                        distance = "Intersects site" if raw_distance < 1 else round(raw_distance, 1)
                        point_gdf = gpd.GeoDataFrame(
                            geometry=[centroid], crs=ANALYSIS_CRS
                        ).to_crs("EPSG:4326")
                        lat = round(point_gdf.geometry.iloc[0].y, 6)
                        lon = round(point_gdf.geometry.iloc[0].x, 6)
                        name = ""
                        if "name" in row and str(row["name"]) not in ["nan", "None", ""]:
                            name = str(row["name"])
                        within_fields = {
                            f"Within {d}m": row.geometry.intersects(
                                buffers[d].geometry.union_all()
                            )
                            for d in buffer_distances
                        }
                        receptor_rows.append({
                            "Type": receptor_type,
                            "Name": name,
                            "Distance to Site (m)": distance,
                            **within_fields,
                            "Latitude": lat,
                            "Longitude": lon,
                        })

                except Exception as e:
                    print(f"Receptor detail export error for {layer_key}: {e}")

            if receptor_rows:
                df_receptors = pd.DataFrame(receptor_rows).sort_values(
                    ["Type", "Distance to Site (m)"]
                )
                df_receptors.to_excel(writer, sheet_name="Receptors Detail", index=False)
                format_sheet(writer.sheets["Receptors Detail"], df_receptors)

        # --------------------------------------------------------
        # SHEET — REFERENCES (last sheet)
        # --------------------------------------------------------
        seen_keys = set()
        ref_rows = []
        for r in results:
            lk = r["layer_key"]
            if lk in seen_keys:
                continue
            seen_keys.add(lk)
            entry = LAYER_REGISTRY.get(lk, {})
            label = entry.get("label", r["layer"])
            code = entry.get("dataset_code", "")
            source = entry.get("source", "")
            theme = entry.get("theme", r.get("theme", ""))
            provider = _get_provider(source) if source else ""
            citation = _get_citation(lk, entry)
            ref_rows.append({
                "Theme": theme,
                "Layer": label,
                "Dataset Code": code,
                "Provider": provider,
                "Citation": citation,
            })

        if ref_rows:
            ref_rows.sort(key=lambda x: (x["Theme"], x["Layer"]))
            df_refs = pd.DataFrame(ref_rows)
            df_refs.to_excel(writer, sheet_name="References", index=False)
            format_sheet(writer.sheets["References"], df_refs)

    print(f"Results exported to: {file_path}")
    return file_path